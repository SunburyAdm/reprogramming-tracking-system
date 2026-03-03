import io
from typing import List, Optional
from uuid import UUID
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func

from app.models import Session, Station, Box, SessionBoxECU, FlashAttempt, User


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STATUS_FILLS = {
    "success": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "failed": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "rework_pending": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "flashing": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "learned": PatternFill(start_color="EDF2F8", end_color="EDF2F8", fill_type="solid"),
    "completed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "blocked": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "in_progress": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "pending": PatternFill(start_color="EDF2F8", end_color="EDF2F8", fill_type="solid"),
    "learning": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
}


def _apply_header_style(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _apply_section_style(cell):
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(vertical="center")


def _write_kv(ws, row: int, key: str, value, col_offset: int = 1) -> int:
    """Write a key-value pair on a single row; returns the row consumed."""
    ws.cell(row=row, column=col_offset, value=key).font = Font(bold=True)
    ws.cell(row=row, column=col_offset + 1, value=value)
    return row + 1


def _auto_width(ws, padding: int = 4):
    """Set column widths based on max content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + padding, 60)


class ReportService:

    # -----------------------------------------------------------------------
    # Box report
    # -----------------------------------------------------------------------

    @staticmethod
    async def generate_box_report(
        db: AsyncSession,
        session_id: UUID,
        box_id: UUID,
    ) -> bytes:
        # ----- fetch data --------------------------------------------------
        session_result = await db.execute(
            select(Session).where(Session.id == session_id)
        )
        session = session_result.scalar_one_or_none()
        if session is None:
            raise ValueError(f"Session {session_id} not found")

        box_result = await db.execute(
            select(Box).where(
                and_(Box.id == box_id, Box.session_id == session_id)
            )
        )
        box = box_result.scalar_one_or_none()
        if box is None:
            raise ValueError(f"Box {box_id} not found")

        station_name = "—"
        if box.assigned_station_id is not None:
            station_result = await db.execute(
                select(Station).where(Station.id == box.assigned_station_id)
            )
            station = station_result.scalar_one_or_none()
            if station is not None:
                station_name = station.name

        ecus_result = await db.execute(
            select(SessionBoxECU).where(
                and_(
                    SessionBoxECU.box_id == box_id,
                    SessionBoxECU.session_id == session_id,
                )
            )
        )
        ecus: List[SessionBoxECU] = ecus_result.scalars().all()

        # Fetch last-user names for each ECU
        user_ids = {e.last_user_id for e in ecus if e.last_user_id is not None}
        user_map: dict = {}
        if user_ids:
            users_result = await db.execute(
                select(User).where(User.id.in_(user_ids))
            )
            for u in users_result.scalars().all():
                user_map[u.id] = u.name or u.email

        # Fetch last attempt notes per ECU
        notes_map: dict = {}
        if ecus:
            ecu_ids = [e.id for e in ecus]
            attempts_result = await db.execute(
                select(FlashAttempt).where(
                    FlashAttempt.ecu_context_id.in_(ecu_ids)
                ).order_by(FlashAttempt.ecu_context_id, FlashAttempt.attempt_no.desc())
            )
            all_attempts: List[FlashAttempt] = attempts_result.scalars().all()
            seen: set = set()
            for a in all_attempts:
                if a.ecu_context_id not in seen:
                    notes_map[a.ecu_context_id] = a.notes or ""
                    seen.add(a.ecu_context_id)

        # ----- build workbook ----------------------------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Box Report"

        row = 1

        # Title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        title_cell = ws.cell(row=row, column=1, value="ECU Reflash — Box Report")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        row += 1

        # ----- header section ----------------------------------------------
        _apply_section_style(ws.cell(row=row, column=1, value="Report Metadata"))
        row += 1

        row = _write_kv(ws, row, "Session Name", session.name)
        row = _write_kv(ws, row, "Target SW Version", session.target_sw_version)
        row = _write_kv(ws, row, "Box Serial", box.box_serial)
        row = _write_kv(ws, row, "Station", station_name)
        row = _write_kv(ws, row, "Box Status", box.status)
        row = _write_kv(
            ws, row, "Frozen At",
            box.frozen_at.strftime("%Y-%m-%d %H:%M:%S UTC") if box.frozen_at else "—",
        )
        row = _write_kv(
            ws, row, "Completed At",
            box.completed_at.strftime("%Y-%m-%d %H:%M:%S UTC") if box.completed_at else "—",
        )
        row = _write_kv(ws, row, "Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"))
        row += 1  # blank row

        # ----- ECU table ---------------------------------------------------
        ecu_headers = [
            "ECU Code",
            "Status",
            "Attempts",
            "Total Time (s)",
            "Last Attempt (s)",
            "Technician",
            "Notes",
        ]
        for col_idx, header in enumerate(ecu_headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            _apply_header_style(cell)
        row += 1

        time_values: List[int] = []
        fail_count = 0
        success_count = 0
        rework_count = 0

        for ecu in ecus:
            tech_name = user_map.get(ecu.last_user_id, "—") if ecu.last_user_id else "—"
            notes_text = notes_map.get(ecu.id, "")
            total_secs = ecu.total_time_seconds or 0
            last_secs = ecu.last_attempt_duration_seconds

            values = [
                ecu.ecu_code,
                ecu.status,
                ecu.attempts or 0,
                total_secs,
                last_secs if last_secs is not None else "—",
                tech_name,
                notes_text,
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")

            status_fill = STATUS_FILLS.get(ecu.status)
            if status_fill:
                for col_idx in range(1, len(ecu_headers) + 1):
                    ws.cell(row=row, column=col_idx).fill = status_fill

            if ecu.status == "success":
                success_count += 1
                if total_secs:
                    time_values.append(total_secs)
            elif ecu.status == "failed":
                fail_count += 1
            elif ecu.status == "rework_pending":
                rework_count += 1

            row += 1

        row += 1  # blank row

        # ----- KPIs --------------------------------------------------------
        _apply_section_style(ws.cell(row=row, column=1, value="KPIs"))
        row += 1

        avg_time = int(sum(time_values) / len(time_values)) if time_values else 0
        max_time = max(time_values) if time_values else 0
        min_time = min(time_values) if time_values else 0

        row = _write_kv(ws, row, "Total ECUs", len(ecus))
        row = _write_kv(ws, row, "Success Count", success_count)
        row = _write_kv(ws, row, "Failed Count", fail_count)
        row = _write_kv(ws, row, "Rework Pending", rework_count)
        row = _write_kv(ws, row, "Avg Flash Time (s)", avg_time)
        row = _write_kv(ws, row, "Max Flash Time (s)", max_time)
        row = _write_kv(ws, row, "Min Flash Time (s)", min_time)

        total = len(ecus)
        fail_rate = round((fail_count / total) * 100, 1) if total else 0.0
        row = _write_kv(ws, row, "Failure Rate (%)", fail_rate)

        _auto_width(ws)
        ws.row_dimensions[1].height = 24

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # -----------------------------------------------------------------------
    # Session report
    # -----------------------------------------------------------------------

    @staticmethod
    async def generate_session_report(
        db: AsyncSession,
        session_id: UUID,
    ) -> bytes:
        # ----- fetch session -----------------------------------------------
        session_result = await db.execute(
            select(Session).where(Session.id == session_id)
        )
        session = session_result.scalar_one_or_none()
        if session is None:
            raise ValueError(f"Session {session_id} not found")

        # ----- fetch boxes -------------------------------------------------
        boxes_result = await db.execute(
            select(Box).where(Box.session_id == session_id)
        )
        boxes: List[Box] = boxes_result.scalars().all()
        box_ids = [b.id for b in boxes]

        # ----- fetch stations ----------------------------------------------
        stations_result = await db.execute(
            select(Station).where(Station.session_id == session_id)
        )
        stations: List[Station] = stations_result.scalars().all()

        # ----- fetch all ECUs ----------------------------------------------
        all_ecus: List[SessionBoxECU] = []
        if box_ids:
            ecus_result = await db.execute(
                select(SessionBoxECU).where(
                    SessionBoxECU.session_id == session_id
                )
            )
            all_ecus = ecus_result.scalars().all()

        # ----- fetch all flash attempts ------------------------------------
        all_attempts: List[FlashAttempt] = []
        if all_ecus:
            ecu_ids = [e.id for e in all_ecus]
            attempts_result = await db.execute(
                select(FlashAttempt).where(
                    FlashAttempt.ecu_context_id.in_(ecu_ids)
                )
            )
            all_attempts = attempts_result.scalars().all()

        # ----- fetch users -------------------------------------------------
        user_ids_all = {e.last_user_id for e in all_ecus if e.last_user_id}
        user_ids_all |= {a.user_id for a in all_attempts if a.user_id}
        user_map: dict = {}
        if user_ids_all:
            users_result = await db.execute(
                select(User).where(User.id.in_(user_ids_all))
            )
            for u in users_result.scalars().all():
                user_map[u.id] = u.name or u.email

        # Index data
        box_map: dict = {b.id: b for b in boxes}
        station_map: dict = {s.id: s for s in stations}
        ecus_by_box: dict = {}
        for ecu in all_ecus:
            ecus_by_box.setdefault(ecu.box_id, []).append(ecu)

        attempts_by_station: dict = {}
        for a in all_attempts:
            if a.station_id:
                attempts_by_station.setdefault(a.station_id, []).append(a)

        # ----- build workbook ----------------------------------------------
        wb = Workbook()

        # ===== Sheet 1: Summary ===========================================
        ws_summary = wb.active
        ws_summary.title = "Summary"

        row = 1
        ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        summary_title = ws_summary.cell(row=row, column=1, value="Session Summary Report")
        summary_title.font = Font(bold=True, size=14)
        summary_title.alignment = Alignment(horizontal="center")
        row += 2

        total_ecus_count = len(all_ecus)
        success_ecus = sum(1 for e in all_ecus if e.status == "success")
        failed_ecus = sum(1 for e in all_ecus if e.status == "failed")
        rework_ecus = sum(1 for e in all_ecus if e.status == "rework_pending")
        completed_boxes = sum(1 for b in boxes if b.status == "completed")
        blocked_boxes = sum(1 for b in boxes if b.status == "blocked")
        in_progress_boxes = sum(1 for b in boxes if b.status == "in_progress")

        success_times = [e.total_time_seconds for e in all_ecus if e.status == "success" and e.total_time_seconds]
        avg_flash = int(sum(success_times) / len(success_times)) if success_times else 0
        fail_rate = round((failed_ecus / total_ecus_count) * 100, 1) if total_ecus_count else 0.0

        kv_rows = [
            ("Session Name", session.name),
            ("Session Status", session.status),
            ("Target SW Version", session.target_sw_version),
            ("Started At", session.started_at.strftime("%Y-%m-%d %H:%M:%S UTC") if session.started_at else "—"),
            ("Closed At", session.closed_at.strftime("%Y-%m-%d %H:%M:%S UTC") if session.closed_at else "—"),
            ("", ""),
            ("Total Boxes", len(boxes)),
            ("Completed Boxes", completed_boxes),
            ("Blocked Boxes", blocked_boxes),
            ("In Progress Boxes", in_progress_boxes),
            ("", ""),
            ("Total ECUs", total_ecus_count),
            ("Success ECUs", success_ecus),
            ("Failed ECUs", failed_ecus),
            ("Rework Pending ECUs", rework_ecus),
            ("", ""),
            ("Overall Failure Rate (%)", fail_rate),
            ("Avg Flash Time (s)", avg_flash),
            ("Total Flash Attempts", len(all_attempts)),
            ("", ""),
            ("Report Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
        ]
        for key, value in kv_rows:
            if key == "":
                row += 1
                continue
            ws_summary.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=value)
            row += 1

        _auto_width(ws_summary)

        # ===== Sheet 2: Boxes =============================================
        ws_boxes = wb.create_sheet("Boxes")
        row = 1
        box_headers = [
            "Box Serial",
            "Status",
            "Station",
            "Expected ECUs",
            "Learned ECUs",
            "Success",
            "Failed",
            "Rework Pending",
            "Inventory Frozen",
            "Frozen At",
            "Completed At",
            "Avg Flash Time (s)",
            "Total Flash Time (s)",
        ]
        for col_idx, header in enumerate(box_headers, start=1):
            _apply_header_style(ws_boxes.cell(row=row, column=col_idx, value=header))
        row += 1

        for box in boxes:
            box_ecus = ecus_by_box.get(box.id, [])
            b_success = sum(1 for e in box_ecus if e.status == "success")
            b_failed = sum(1 for e in box_ecus if e.status == "failed")
            b_rework = sum(1 for e in box_ecus if e.status == "rework_pending")
            b_times = [e.total_time_seconds for e in box_ecus if e.total_time_seconds]
            b_avg = int(sum(b_times) / len(b_times)) if b_times else 0
            b_total_time = sum(b_times)
            stn = station_map.get(box.assigned_station_id)
            stn_name = stn.name if stn else "—"

            row_values = [
                box.box_serial,
                box.status,
                stn_name,
                box.expected_ecu_count if box.expected_ecu_count is not None else "—",
                box.learned_count or 0,
                b_success,
                b_failed,
                b_rework,
                "Yes" if box.inventory_frozen else "No",
                box.frozen_at.strftime("%Y-%m-%d %H:%M:%S") if box.frozen_at else "—",
                box.completed_at.strftime("%Y-%m-%d %H:%M:%S") if box.completed_at else "—",
                b_avg,
                b_total_time,
            ]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws_boxes.cell(row=row, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")

            status_fill = STATUS_FILLS.get(box.status)
            if status_fill:
                for col_idx in range(1, len(box_headers) + 1):
                    ws_boxes.cell(row=row, column=col_idx).fill = status_fill

            row += 1

        _auto_width(ws_boxes)

        # ===== Sheet 3: Stations ==========================================
        ws_stations = wb.create_sheet("Stations")
        row = 1
        station_headers = [
            "Station Name",
            "Total Attempts",
            "Success",
            "Failed",
            "Success Rate (%)",
            "Total Time (s)",
            "Avg Time per Attempt (s)",
            "ECUs / Hour",
        ]
        for col_idx, header in enumerate(station_headers, start=1):
            _apply_header_style(ws_stations.cell(row=row, column=col_idx, value=header))
        row += 1

        for station in stations:
            s_attempts = attempts_by_station.get(station.id, [])
            s_total = len(s_attempts)
            s_success = sum(1 for a in s_attempts if a.result == "success")
            s_failed = sum(1 for a in s_attempts if a.result == "failed")
            s_success_rate = round((s_success / s_total) * 100, 1) if s_total else 0.0
            s_durations = [a.duration_seconds for a in s_attempts if a.duration_seconds is not None]
            s_total_time = sum(s_durations)
            s_avg_time = int(s_total_time / len(s_durations)) if s_durations else 0
            ecus_per_hour = round((s_success / (s_total_time / 3600)), 1) if s_total_time else 0.0

            row_values = [
                station.name,
                s_total,
                s_success,
                s_failed,
                s_success_rate,
                s_total_time,
                s_avg_time,
                ecus_per_hour,
            ]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws_stations.cell(row=row, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
            row += 1

        _auto_width(ws_stations)

        # ===== Sheet 4: ECUs ==============================================
        ws_ecus = wb.create_sheet("ECUs")
        row = 1
        ecu_headers = [
            "ECU Code",
            "Box Serial",
            "Station",
            "Status",
            "Attempts",
            "Total Time (s)",
            "Last Attempt (s)",
            "Last Technician",
        ]
        for col_idx, header in enumerate(ecu_headers, start=1):
            _apply_header_style(ws_ecus.cell(row=row, column=col_idx, value=header))
        row += 1

        for ecu in all_ecus:
            box_obj = box_map.get(ecu.box_id)
            box_serial = box_obj.box_serial if box_obj else "—"
            stn_id = box_obj.assigned_station_id if box_obj else None
            stn_name_ecu = station_map[stn_id].name if stn_id and stn_id in station_map else "—"
            tech = user_map.get(ecu.last_user_id, "—") if ecu.last_user_id else "—"

            row_values = [
                ecu.ecu_code,
                box_serial,
                stn_name_ecu,
                ecu.status,
                ecu.attempts or 0,
                ecu.total_time_seconds or 0,
                ecu.last_attempt_duration_seconds if ecu.last_attempt_duration_seconds is not None else "—",
                tech,
            ]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws_ecus.cell(row=row, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")

            status_fill = STATUS_FILLS.get(ecu.status)
            if status_fill:
                for col_idx in range(1, len(ecu_headers) + 1):
                    ws_ecus.cell(row=row, column=col_idx).fill = status_fill

            row += 1

        _auto_width(ws_ecus)

        # ----- serialise ---------------------------------------------------
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
